from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

FILE_NAME = "registrations.xlsx"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['POST'])
def register():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    game = request.form['game']
    gender = request.form['gender']

    # If file doesn't exist, create it
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone", "Game", "Gender"])
        wb.save(FILE_NAME)

    # Load workbook and append data
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([name, email, phone, game, gender])
    wb.save(FILE_NAME)

    return "<h2 style='color:green;text-align:center;'>Registration Successful!</h2><a href='/'>Go Back</a>"

if __name__ == '__main__':
    app.run(debug=True)
